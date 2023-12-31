from fpdf import FPDF
import pandas as pd
import glob
import openpyxl
from pathlib import Path

filepaths = glob.glob("invoices/*.xlsx")

for filepath in filepaths:
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.add_page()
    filename = Path(filepath).stem
    invoiceNo = filename.split("-")[0]
    date = filename.split("-")[1]
    pdf.set_font(family="Times", size=16, style="B")
    pdf.cell(w=50, h=8, txt=f"Invoice nr. {invoiceNo}", ln = 1)

    pdf.set_font(family="Times", size=16, style="B")
    pdf.cell(w=50, h=8, txt=f"Date. {date}",ln = 1)
    df = pd.read_excel(filepath, sheet_name="Sheet 1")

    colums = list(df.columns)
    colums = [item.replace("_", " ").title() for item in colums]
    pdf.set_font(family="Times", size=10, style="B")
    pdf.set_text_color(80, 80, 80)
    pdf.cell(w=30, h=8, txt=colums[0], border=1)
    pdf.cell(w=70, h=8, txt=colums[1], border=1)
    pdf.cell(w=30, h=8, txt=colums[2], border=1)
    pdf.cell(w=30, h=8, txt=colums[3], border=1)
    pdf.cell(w=30, h=8, txt=colums[4], border=1, ln=1)

    for index, row in df.iterrows():
        pdf.set_font(family="Times", size=10)
        pdf.set_text_color(80,80,80)
        pdf.cell(w=30, h=8, txt=str(row["product_id"]), border=1)
        pdf.cell(w=70, h=8, txt=str(row["product_name"]), border=1)
        pdf.cell(w=30, h=8, txt=str(row["amount_purchased"]), border=1)
        pdf.cell(w=30, h=8, txt=str(row["price_per_unit"]), border=1)
        pdf.cell(w=30, h=8, txt=str(row["total_price"]), border=1, ln=1)

    total_sum = df["total_price"].sum()
    pdf.set_font(family="Times", size=10)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(w=30, h=8, txt="", border=1)
    pdf.cell(w=70, h=8, txt="", border=1)
    pdf.cell(w=30, h=8, txt="", border=1)
    pdf.cell(w=30, h=8, txt="", border=1)
    pdf.cell(w=30, h=8, txt=str(total_sum), border=1, ln=1)

    pdf.set_font(family="Times", size=10)
    pdf.cell(w=30, h=8, txt=f"The total sum is {str(total_sum)}", ln=1)

    pdf.set_font(family="Times", size=10)
    pdf.cell(w=30, h=8, txt=f"PythonHow")
    pdf.image("pythonhow.png", w=10)


    pdf.output(f"PDFs/{filename}.pdf")



